"""
社会保険料一括計算ツール（令和8年度）
対象：社労士・HR担当・中小企業総務

【構成】
  無料機能：従業員負担の計算・表示・Excel出力
  有料機能：事業主負担の計算・表示（パスワード認証後に解放）
"""

import io
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from calculator import (
    fetch_rates,
    calc_dataframe,
    PREFS,
    KOYO_RATES_WORKER,
)

# ============================================================
# ページ設定
# ============================================================
st.set_page_config(
    page_title="社会保険料一括計算（令和8年度）",
    page_icon="📊",
    layout="wide",
)

# ============================================================
# 料率データ読み込み（1日キャッシュ）
# ============================================================
@st.cache_data(ttl=3600 * 24, show_spinner=False)
def load_rates():
    return fetch_rates()


# ============================================================
# テーブル表示（共通）
# ============================================================
def show_table(result: pd.DataFrame, cols: list):
    fmt = {c: "{:,.0f}" for c in cols if c != "氏名"}
    st.dataframe(
        result[cols].style.format(fmt),
        use_container_width=True,
        hide_index=True,
    )


# ============================================================
# Excel出力
# ============================================================
def to_excel(df_apr: pd.DataFrame, df_may: pd.DataFrame,
             cols: list) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_apr[cols].to_excel(writer, index=False, sheet_name="4月支払い分")
        df_may[cols].to_excel(writer, index=False, sheet_name="5月支払い分以降")

    buf.seek(0)
    wb = load_workbook(buf)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
                    cell.alignment     = Alignment(horizontal="right")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ============================================================
# パスワード認証UI（事業主負担の解放）
# ============================================================
def show_password_form():
    """
    パスワード入力フォームを表示する。
    認証成功時は session_state["employer_unlocked"] = True にして rerun。
    """
    try:
        correct_pw = st.secrets["app_password"]
    except Exception:
        # Secrets 未設定（ローカル開発時）はそのまま解放
        st.session_state["employer_unlocked"] = True
        return

    with st.expander("🔑 事業主負担を表示する（有料機能）"):
        st.caption("パスワードをお持ちの方は入力してください。")
        pw_input = st.text_input("パスワード", type="password", key="pw_input")
        if st.button("認証する", key="auth_btn"):
            if pw_input == correct_pw:
                st.session_state["employer_unlocked"] = True
                st.rerun()
            else:
                st.error("パスワードが違います")


# ============================================================
# メイン画面
# ============================================================
st.title("📊 社会保険料一括計算ツール")
st.caption("令和8年度 協会けんぽ対応｜複数名同時計算・Excel出力")

# 料率読み込み
with st.spinner("協会けんぽから料率データを取得中..."):
    try:
        rates = load_rates()
        st.success(f"✅ 令和8年度 料率データ読み込み完了（{len(rates)} 都道府県）")
    except Exception as e:
        st.error(f"料率の取得に失敗しました: {e}")
        st.stop()

st.divider()

# ============================================================
# 従業員データ入力
# ============================================================
st.subheader("👥 従業員データ入力")
st.caption("行を追加して複数名を一括入力できます。標準報酬月額を入力してください（総支給額ではありません）。")

default_df = pd.DataFrame({
    "氏名":         ["（例）田中 太郎", "（例）鈴木 花子"],
    "年齢":         [45, 38],
    "都道府県":     ["東京", "東京"],
    "標準報酬月額": [300_000, 250_000],
    "業種":         ["一般", "一般"],
})

edited_df = st.data_editor(
    default_df,
    num_rows="dynamic",
    column_config={
        "氏名": st.column_config.TextColumn("氏名", required=True),
        "年齢": st.column_config.NumberColumn(
            "年齢", min_value=15, max_value=100, step=1, format="%d歳"
        ),
        "都道府県": st.column_config.SelectboxColumn(
            "都道府県（事業所所在地）", options=PREFS, required=True
        ),
        "標準報酬月額": st.column_config.NumberColumn(
            "標準報酬月額（円）", min_value=0, step=1000, format="%d円"
        ),
        "業種": st.column_config.SelectboxColumn(
            "業種（雇用保険料率区分）",
            options=list(KOYO_RATES_WORKER.keys()),
            required=True,
        ),
    },
    use_container_width=True,
    hide_index=True,
)

st.divider()

# ============================================================
# 「計算する」ボタン
# 押されたときだけ計算してセッションに保存する
# ============================================================
if st.button("　計算する　", type="primary", use_container_width=True):

    df_clean = edited_df.dropna(subset=["氏名", "標準報酬月額", "年齢", "都道府県"])
    df_clean = df_clean[df_clean["標準報酬月額"] > 0]

    if df_clean.empty:
        st.warning("有効な従業員データを1名以上入力してください。")
        st.stop()

    # 計算結果をセッションに保存
    # → パスワード認証で rerun されても結果が消えない
    st.session_state["result_apr"] = calc_dataframe(df_clean, rates, shien_kin=False)
    st.session_state["result_may"] = calc_dataframe(df_clean, rates, shien_kin=True)

# ============================================================
# 計算結果の表示
# ボタンブロックの外に出すことで rerun 後も表示を維持する
# ============================================================
if "result_apr" in st.session_state:

    result_apr = st.session_state["result_apr"]
    result_may = st.session_state["result_may"]

    emp_cols = ["氏名", "標準報酬月額",
                "健康保険料(本人)", "子育て支援金(本人)",
                "厚生年金料(本人)", "雇用保険料(本人)",
                "控除合計(本人)", "手取り概算"]

    # ===== 従業員負担（無料） =====
    st.subheader("📋 計算結果①：従業員負担（控除額・手取り）")

    tab1, tab2 = st.tabs(
        ["4月支払い分（子育て支援金 なし）", "5月支払い分以降（子育て支援金 あり）"]
    )
    with tab1:
        show_table(result_apr, emp_cols)
        c1, c2 = st.columns(2)
        c1.metric("控除合計（全員）",       f"{int(result_apr['控除合計(本人)'].sum()):,} 円")
        c2.metric("手取り概算合計（全員）", f"{int(result_apr['手取り概算'].sum()):,} 円")
    with tab2:
        show_table(result_may, emp_cols)
        c1, c2 = st.columns(2)
        c1.metric("控除合計（全員）",       f"{int(result_may['控除合計(本人)'].sum()):,} 円")
        c2.metric("手取り概算合計（全員）", f"{int(result_may['手取り概算'].sum()):,} 円")

    st.divider()

    # ===== 事業主負担（パスワード認証後） =====
    if st.session_state.get("employer_unlocked", False):

        sha_cols = ["氏名", "標準報酬月額",
                    "健康保険料(会社)", "子育て支援金(会社)",
                    "厚生年金料(会社)", "雇用保険料(会社)",
                    "会社負担合計", "人件費総額"]

        st.subheader("📋 計算結果②：事業主負担・人件費総額")
        st.caption("⚠️ 事業主の雇用保険料率は毎年変更される可能性があります。厚労省公式資料でご確認ください。")

        tab3, tab4 = st.tabs(
            ["4月支払い分（子育て支援金 なし）", "5月支払い分以降（子育て支援金 あり）"]
        )
        with tab3:
            show_table(result_apr, sha_cols)
            c1, c2 = st.columns(2)
            c1.metric("会社負担合計（全員）", f"{int(result_apr['会社負担合計'].sum()):,} 円")
            c2.metric("人件費総額（全員）",   f"{int(result_apr['人件費総額'].sum()):,} 円")
        with tab4:
            show_table(result_may, sha_cols)
            c1, c2 = st.columns(2)
            c1.metric("会社負担合計（全員）", f"{int(result_may['会社負担合計'].sum()):,} 円")
            c2.metric("人件費総額（全員）",   f"{int(result_may['人件費総額'].sum()):,} 円")

        # Excel（完全版）
        st.divider()
        all_cols = emp_cols + ["健康保険料(会社)", "子育て支援金(会社)",
                               "厚生年金料(会社)", "雇用保険料(会社)",
                               "会社負担合計", "人件費総額"]
        excel_bytes = to_excel(result_apr, result_may, all_cols)
        st.download_button(
            label="📥 Excelダウンロード（完全版：従業員負担＋事業主負担）",
            data=excel_bytes,
            file_name="社会保険料計算_令和8年度_完全版.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    else:
        # 未認証：パスワードフォームを表示
        show_password_form()

        # Excel（従業員のみ）
        st.divider()
        excel_bytes = to_excel(result_apr, result_may, emp_cols)
        st.download_button(
            label="📥 Excelダウンロード（従業員負担のみ）",
            data=excel_bytes,
            file_name="社会保険料計算_令和8年度.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

# ============================================================
# フッター注記
# ============================================================
st.divider()
with st.expander("ℹ️ 計算の前提・注意事項"):
    st.markdown("""
    - **対象**：全国健康保険協会（協会けんぽ）加入事業所
    - **健康保険料**：標準報酬月額 × 都道府県別料率 ÷ 2（四捨五入）
    - **介護保険**：年齢40歳以上の場合、介護保険料率を含む料率を適用
    - **子ども・子育て支援金**：標準報酬月額 × 0.23% ÷ 2（5月支払い分以降）
    - **厚生年金保険料**：標準報酬月額 × 18.3% ÷ 2（四捨五入）
    - **雇用保険料（本人）**：標準報酬月額 × 業種別料率（切り捨て）
    - **雇用保険料（会社）**：毎年3月に自動更新（厚労省公式PDFより取得）
    - **端数処理**：健保・年金は50銭以上切り上げ、雇用保険は切り捨て
    - 本ツールの計算結果は参考値です。実務では必ず保険料額表でご確認ください。
    """)
